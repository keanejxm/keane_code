#!/usr/bin/python3
# -*- coding:utf-8 -*-
"""
:author: keane
:file  doc_convert_pdf.py
:time  2023/9/14 10:50
:desc  
"""
import sys
import fitz
import pandas as pd
from openpyxl import load_workbook

from win32com.client import Dispatch, constants, gencache, DispatchEx


# 使用pywin32转pdf
def doc(filename, exportfile):
    wordapp = Dispatch("Word.Application")
    # wordapp = Dispatch("kwps.Application")
    try:
        # gencache.EnsureModule('{00020905-0000-0000-C000-000000000046}', 0, 8, 4)
        doc = wordapp.Documents.Open(filename)
        doc.ExportAsFixedFormat(exportfile, 17)
        doc.Close()
        wordapp.Quit()
    except Exception as e:
        print(f"转换异常，异常为{e}")
    finally:
        wordapp.Quit()


def xls(filename, exportfile):
    """
    在微软的excel激活的情况或者使用wps时都可转为pdf
    当excel未激活是转pdf会失败
    :param filename:
    :param exportfile:
    :return:
    """
    xlapp = DispatchEx("Excel.Application")
    try:
        xlapp.Visible = False
        xlapp.DisplayAlerts = False
        books = xlapp.Workbooks.Open(filename, False)
        books.ExportAsFixedFormat(0, exportfile)
        books.Close()
        xlapp.Quit()
    except Exception as e:
        print(f"转换异常，异常为{e}")
    finally:
        xlapp.Quit()

file_name = "E:\keane_python\github\keane_code\办公自动化\辅助账余额取数-20230804-不限制科目.xlsx"
html_name = "E:\keane_python\github\keane_code\办公自动化\信披取数规则 0822.html"
export_file = "E:\keane_python\github\keane_code\办公自动化\辅助账余额取数-20230804-不限制科目.pdf"
from openpyxl import load_workbook
from reportlab.pdfgen import canvas

# 加载Excel文件
wb = load_workbook(filename=file_name)

# 选择要转换的工作表
ws = wb.active

# 创建PDF文件
pdf_file = canvas.Canvas(export_file)

# 循环遍历每一行并写入PDF文件
for row in ws.iter_rows(min_row=2):
    for cell in row:
        pdf_file.drawString(cell.column_dimensions.width * 0.5 + 10, cell.row_dimensions.height * 0.5 + 10, str(cell.value))

# 保存PDF文件
pdf_file.save()

# print(sys.platform)
# # file_name = "E:\keane_python\github\keane_code\办公自动化\开放式基金账户业务申请表(产品).docx"
# # export_file = "E:\keane_python\github\keane_code\办公自动化\开放式基金账户业务申请表(产品).pdf"
# # doc(file_name, export_file)
# file_name = "E:\keane_python\github\keane_code\办公自动化\财务账套.xls"
# export_file = "E:\keane_python\github\keane_code\办公自动化\财务账套.pdf"
# xls(file_name, export_file)
# # excel_data = pd.read_excel(file_name)
